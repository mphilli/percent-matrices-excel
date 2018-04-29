# Create a percent matrix completeness report of metadata, where each reported row corresponds to a particular column

from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell
import xlsxwriter


class ExcelFileSplitter:

    def __init__(self, file, column, value="Sheet1"):
        """Reports on the completeness of a single Excel file, grouped by a specified column"""
        if not file.endswith(".xlsx"):
            file += ".xlsx"
        self.file = file
        self.value = value
        self.column = column
        self.row_keys = self.get_row_keys()
        self.col_names = self.get_col_names()
        self.col_names.remove(self.column)
        self.report = self.percent_info_by_row_name()

    def get_row_keys(self, value="Sheet1"):
        """Retrieve a list of information from each row,
        associating each piece of information with the appropriate column name."""
        wb = load_workbook(filename=self.file)  # create the workbook object based on the Excel file
        sheet = wb[value].values  # Change sheet name to desired output
        values = {}
        row_keys = []
        for i, item in enumerate(sheet):
            this_row = {}
            if i == 0:
                for j in range(len(list(item))):
                    values[j] = item[j]
            else:
                row = list(item)
                for k, r in enumerate(row):
                    this_row[values[k]] = r
            if this_row:
                row_keys.append(this_row)
        return row_keys

    def get_col_names(self):
        """get the names of the columns in this Excel file worksheet"""
        wb = load_workbook(filename=self.file)
        sheet = wb[self.value].values  # Change sheet name to desired output
        names = []
        for i, item in enumerate(sheet):
            if i == 0:
                names = list(item)
            else:
                break
        return names

    def percent_info_by_row_name(self):
        """collects the percentage information for a single spreadsheet, based on a specified column name"""
        field_cache = {}  # each column and its info
        field_count = {}  # the number of rows of data associated with each collection
        for row in self.row_keys:  # for each row from the metadata report
            for col in row:  # for each column (metadata field) from the row
                if col == self.column:  # if the metadata field is the "issue"
                    if row[col] not in field_cache and row[col] not in field_count:
                        # if the issue isn't already in the cache or count dictionaries:
                        field_count[row[col]] = 1  # set the initial field_count to 1 for this issue
                        field_cache[row[col]] = row.copy()  # make a copy of the current row with this issue
                        for item in field_cache[row[col]]:
                            if item == self.column:
                                del field_cache[row[col]][item]
                                break
                        print(field_cache)
                        for item in field_cache[row[col]]:  # for each metadata element in the copied row
                            if not field_cache[row[col]][item] or \
                                   field_cache[row[col]][item] == "":  # if it's blank, set initial value to 0
                                field_cache[row[col]][item] = 0
                            else:
                                field_cache[row[col]][item] = 1    # otherwise, set initial value to 1

                    else:
                        # if we've already encountered this collection
                        field_count[row[col]] += 1  # increment the field_count for this issue
                        for column in row:
                            for item in field_cache[row[col]]:
                                if item == column:
                                    if row[column] != "" and row[column]:
                                        field_cache[row[col]][item] += 1
        for i in field_cache:
            for j in field_count:
                if i == j:
                    for item in field_cache[i]:
                        field_cache[i][item] = str("=" + str(field_cache[i][item])
                                                   + "/" + str(field_count[j])
                                                   + "*100")  # create Excel formula value
        return field_cache

    def generate_csv_matrix(self, name=""):
        if not name.endswith(".xlsx"):
            name += ".xlsx"
        workbook = xlsxwriter.Workbook(name)
        worksheet = workbook.add_worksheet(name=self.value)
        excel_data = [[""] + self.col_names]
        """function for generating the final CSV file with the percent matrix of SURFACE information"""

        for iss in self.report:  # for each SURFACE issue...
            row_print = [iss]  # initialize the row to be printed corresponding to each issue
            for col in self.col_names:  # for each metadata element...
                for item in self.report[iss]:  # for each metadata element associated with the issues
                    if col == item:  # if the name of the metadata element matches the current column name
                        row_print.append(self.report[iss][item])
                        # append the value inside the issues dictionary to the row
                        # (e.g., =1/19*100)
            excel_data.append(row_print)
        wb_format = workbook.add_format()
        wb_format.set_font_size(12)
        for row, row_value in enumerate(excel_data):
            worksheet.write_row(row, 0, row_value, wb_format)

        peak = xl_rowcol_to_cell(len(excel_data), len(excel_data[0]))
        worksheet.conditional_format("A1:" + peak, {'type': '3_color_scale'})


if __name__ == "__main__":
    # run demo on color data
    efs = ExcelFileSplitter(file="demo/sample_file.xlsx", column='collection')
    efs.generate_csv_matrix(name="demo/output/SampleFile.xlsx")
