# A Python program for reporting the completeness of data in a spreadsheet or set of spreadsheets
# as a percentage matrix. It is currently required that the sheet name be the same in each Excel file.

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join, dirname, realpath


class ExcelFileCombiner:

    def __init__(self,
                 directory="files/in",
                 sheet_name="Sheet1",
                 output="files/out/efc_output.xlsx"):
        """Combine all data files into a single completeness report"""
        self.value = sheet_name
        self.dir_path = dirname(realpath(__file__))
        self._directory = directory
        self._out_file = output
        self.excel_files = self.get_excel_files()
        self.col_names = self.get_all_cols(value=self.value)

    @property
    def _directory(self):
        return self.directory

    @_directory.setter
    def _directory(self, value):
        if value.endswith("/"):
            self.directory = join(self.dir_path, value)
        else:
            self.directory = join(self.dir_path, value + "/")

    @property
    def _out_file(self):
        return self._out_file

    @_out_file.setter
    def _out_file(self, out_val):
        if not out_val.endswith(".xlsx"):
            out_val += ".xlsx"
        self.out_file = join(self.dir_path, out_val)

    def get_percent_matrix(self):
        self.create_percent_matrix(report=self.get_percent_info())

    def get_excel_files(self):
        """Crawl the directory and read each Excel file"""
        return [f for f in listdir(self.directory)  # f = file
                if isfile(join(self.directory, f)) and f.endswith(".xlsx")]

    def get_all_cols(self, value="Sheet1"):
        """Print the name of every column name from every spreadsheet.
           This creates a masterlist of all column names in the directory."""
        col_names = []
        for file in self.get_excel_files():
            for name in self.get_col_names(file, value=value):
                if name not in col_names and name:
                    col_names.append(name)
        return col_names

    def get_col_names(self, file, value):
        """get the names of the columns in this Excel file worksheet"""
        wb = load_workbook(filename=self.directory + str(file))
        sheet = wb[value].values  # Change sheet name to desired output
        names = []
        for i, item in enumerate(sheet):
            if i == 0:
                names = list(item)
            else:
                break
        return names

    def get_file_report(self, file, value):
        """produces the percentage reports for each metadata field for a given collection (e.g., =20/30*100)"""
        wb = load_workbook(filename=self.directory + str(file))
        values = {}
        col_names = self.get_col_names(file, value=value)
        total = len(list(wb[value].values)) - 1
        sheet = wb[value].values
        for i, row in enumerate(sheet):
            if i == 0:
                pass  # skip column names
            else:
                for j, value in enumerate(row):
                    if value or value == 0:
                        if col_names[j] not in values:
                            values[col_names[j]] = 1
                        else:
                            values[col_names[j]] = values[col_names[j]] + 1
                    else:
                        if col_names[j] not in values:
                            values[col_names[j]] = 0
        for value in values:
            if values[value] == 0:
                values[value] = "=0"
            else:
                values[value] = "=" + str(values[value]) + "/" + str(total) + "*100"
        return values

    @staticmethod
    def parse_formula(formula):
        if "/" in formula:
            numerator = float(formula.split("/")[0][1:])
            denominator = float(formula.split("/")[1].split("*")[0])
            return (numerator / denominator) * 100
        else:
            return 0

    def get_percent_info(self, threshold=-1):
        report = []
        for file in self.excel_files:
            results = self.get_file_report(file, value=self.value)
            for r in results:
                if 101 > self.parse_formula(results[r]) > float(threshold):
                    report.append(file.replace(".xlsx", "") + " - " + str(r) + ": " + str(results[r]))
        return report

    def create_percent_matrix(self, report):
        """uses the completeness reports from get_percent_missing results to produce a percent matrix file"""
        # write initial row of column names for this sheet
        workbook = xlsxwriter.Workbook(self.out_file)
        worksheet = workbook.add_worksheet(name=self.value)
        # flatten results of the report into a dictionary object
        rep_dict = {}
        for r in report:
            # {collection: [(col_name, percentage), ...]}
            collection = r.split(" - ")[0]
            percent_pair = r.split(" - ")[1]
            column_name = percent_pair.split(":")[0]
            percentage = percent_pair.split(": ")[1]

            if collection not in rep_dict:
                rep_dict[collection] = [(column_name, percentage)]
            else:
                rep_dict[collection].append((column_name, percentage))
        excel_data = [[""] + self.col_names]
        for file in self.excel_files:
            row_data = []
            fname = file.replace(".xlsx", "")
            row_data.append(fname)
            for rep in rep_dict:
                if rep == fname:
                    for col in self.col_names:
                        this_value = None
                        for tup in rep_dict[rep]:
                            if tup[0] == col:
                                this_value = str(tup[1])
                                break
                        if this_value:
                            row_data.append(this_value)
                        else:
                            row_data.append("")
            excel_data.append(row_data)

        wb_format = workbook.add_format()
        wb_format.set_font_size(12)
        for row, row_value in enumerate(excel_data):
            worksheet.write_row(row, 0, row_value, wb_format)

        peak = xl_rowcol_to_cell(len(excel_data), len(excel_data[0]))
        worksheet.conditional_format("A1:" + peak, {'type': '3_color_scale'})
        workbook.close()


if __name__ == "__main__":
    # run demo on color sample_set
    efc = ExcelFileCombiner(directory="demo/sample_set", output="demo/output/ColorPercentMatrix")
    efc.get_percent_matrix()
